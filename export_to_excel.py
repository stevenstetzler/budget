import argparse
import datetime
import os

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, Side
from openpyxl.utils import get_column_letter

# Number formats
# Sections: positive, negative (parentheses), zero, text
ACCOUNTING_FORMAT = '_($* #,##0.00_);_($* (#,##0.00);_($* "-"??_);_(@_)'
PERCENT_FORMAT = "0%"

# Expense category priority order (after "Income").  Categories not listed
# here are appended in first-seen insertion order.
PRIORITY_CATS = ["withholding", "taxes", "savings", "rent", "utilities", "car"]

# Border style used throughout
THIN = Side(style="thin")

# Column width constants
SUMMARY_COLUMN_WIDTHS = {
    "A": 18, "B": 15, "C": 10,
    "D": 18, "E": 15, "F": 10, "G": 15, "H": 10,
}
RECEIPTS_DESC_COL_WIDTH = 25
RECEIPTS_AMT_COL_WIDTH = 15


def _update_cell_border(cell, left=None, right=None, top=None, bottom=None):
    """Update specific sides of a cell's border, preserving all other sides."""
    b = cell.border
    cell.border = Border(
        left=left if left is not None else b.left,
        right=right if right is not None else b.right,
        top=top if top is not None else b.top,
        bottom=bottom if bottom is not None else b.bottom,
    )


def _to_bool(val):
    """Coerce truthy/falsy values (bool, str, int) to Python bool."""
    if isinstance(val, bool):
        return val
    if isinstance(val, str):
        return val.strip().lower() == "true"
    return bool(val)


def _write_summary_sheet(ws, most_recent_month, most_recent_year):
    """
    Write the Summary tab layout.

    Layout
    ------
    A1:A2   "Date"              (merged, bold, centered)
    B1:C2   <selected month>    (merged, date value formatted as "MMM YYYY")
    D1:D2   "Cash Flow"         (merged, bold, centered)
    E1      "Amount"            (bold)
    F1      "Fraction"          (bold)
    G1      "Budget"            (bold)
    H1      "Fraction"          (bold)
    E2      =$B$4-E4
    F2      =IFERROR(INDIRECT(ADDRESS(ROW(),COLUMN()-1))/$B$4,"")
    G2      =$B$4-G4
    H2      =IFERROR(INDIRECT(ADDRESS(ROW(),COLUMN()-1))/$B$4,"")
    A3:C3   "Income"            (merged, bold, centered)
    D3:H3   "Expenses"          (merged, bold, centered)
    A4      "Total"             (bold)
    B4      =SUM(B6:B1048576)
    D4      "Total"             (bold)
    E4      =SUM(E6:E102)
    F4      =E4/$B$4
    G4      =SUM(G6:G102)
    H4      =G4/$B$4
    A5-H5   column headers      (all bold)
    A6:H105 per-row IFERROR/ISBLANK/INDIRECT formulas referencing the selected
            receipts sheet; ISBLANK prevents empty cells from showing 0
    """
    bold = Font(bold=True)
    center = Alignment(horizontal="center", vertical="center")

    # --- Row 1 / Row 2 merged headers ---

    # A1:A2 "Date"
    ws.merge_cells(start_row=1, start_column=1, end_row=2, end_column=1)
    ws.cell(row=1, column=1).value = "Date"
    ws.cell(row=1, column=1).font = bold
    ws.cell(row=1, column=1).alignment = center

    # B1:C2 — date value for the selected month
    ws.merge_cells(start_row=1, start_column=2, end_row=2, end_column=3)
    ws.cell(row=1, column=2).value = datetime.date(most_recent_year, most_recent_month, 1)
    ws.cell(row=1, column=2).number_format = "MMM YYYY"
    ws.cell(row=1, column=2).alignment = center

    # D1:D2 "Cash Flow"
    ws.merge_cells(start_row=1, start_column=4, end_row=2, end_column=4)
    ws.cell(row=1, column=4).value = "Cash Flow"
    ws.cell(row=1, column=4).font = bold
    ws.cell(row=1, column=4).alignment = center

    # E1, F1, G1, H1 — column headers
    for col, text in [(5, "Amount"), (6, "Fraction"), (7, "Budget"), (8, "Fraction")]:
        ws.cell(row=1, column=col).value = text
        ws.cell(row=1, column=col).font = bold
        ws.cell(row=1, column=col).alignment = center

    # Row 2 — cash flow formulas
    ws.cell(row=2, column=5).value = "=$B$4-E4"
    ws.cell(row=2, column=5).number_format = ACCOUNTING_FORMAT
    ws.cell(row=2, column=6).value = (
        '=IFERROR(INDIRECT(ADDRESS(ROW(),COLUMN()-1))/$B$4,"")'
    )
    ws.cell(row=2, column=6).number_format = PERCENT_FORMAT
    ws.cell(row=2, column=7).value = "=$B$4-G4"
    ws.cell(row=2, column=7).number_format = ACCOUNTING_FORMAT
    ws.cell(row=2, column=8).value = (
        '=IFERROR(INDIRECT(ADDRESS(ROW(),COLUMN()-1))/$B$4,"")'
    )
    ws.cell(row=2, column=8).number_format = PERCENT_FORMAT

    # --- Row 3 — section labels ---
    ws.merge_cells(start_row=3, start_column=1, end_row=3, end_column=3)
    ws.cell(row=3, column=1).value = "Income"
    ws.cell(row=3, column=1).font = bold
    ws.cell(row=3, column=1).alignment = center

    ws.merge_cells(start_row=3, start_column=4, end_row=3, end_column=8)
    ws.cell(row=3, column=4).value = "Expenses"
    ws.cell(row=3, column=4).font = bold
    ws.cell(row=3, column=4).alignment = center

    # --- Row 4 — totals ---
    ws.cell(row=4, column=1).value = "Total"
    ws.cell(row=4, column=1).font = bold
    ws.cell(row=4, column=1).alignment = center
    ws.cell(row=4, column=2).value = "=SUM(B6:B1048576)"
    ws.cell(row=4, column=2).number_format = ACCOUNTING_FORMAT
    ws.cell(row=4, column=4).value = "Total"
    ws.cell(row=4, column=4).font = bold
    ws.cell(row=4, column=4).alignment = center
    ws.cell(row=4, column=5).value = "=SUM(E6:E102)"
    ws.cell(row=4, column=5).number_format = ACCOUNTING_FORMAT
    ws.cell(row=4, column=6).value = "=E4/$B$4"
    ws.cell(row=4, column=6).number_format = PERCENT_FORMAT
    ws.cell(row=4, column=7).value = "=SUM(G6:G102)"
    ws.cell(row=4, column=7).number_format = ACCOUNTING_FORMAT
    ws.cell(row=4, column=8).value = "=G4/$B$4"
    ws.cell(row=4, column=8).number_format = PERCENT_FORMAT

    # --- Row 5 — column headers ---
    for col, text in [
        (1, "Source"), (2, "Amount"), (3, "Fraction"),
        (4, "Item"), (5, "Amount"), (6, "Fraction"), (7, "Budget"), (8, "Fraction"),
    ]:
        ws.cell(row=5, column=col).value = text
        ws.cell(row=5, column=col).font = bold
        ws.cell(row=5, column=col).alignment = center

    # --- Rows 6–105 — per-row formulas ---
    # Income (cols A–C): individual income transactions from the receipts sheet
    # Expenses (cols D–H): one row per expense category (name, total, %, budget, %)
    # ISBLANK is used so that empty cells in the receipts sheet show "" not 0.
    _ref = (
        "\"'\"&MONTH($B$1)&\" \"&YEAR($B$1)&\" Receipts'!\""
    )
    for row in range(6, 106):
        ws.cell(row=row, column=1).value = (
            f'=IFERROR(IF(ISBLANK(INDIRECT({_ref}&ADDRESS(ROW(),1))),"",INDIRECT({_ref}&ADDRESS(ROW(),1))),"")'
        )
        ws.cell(row=row, column=2).value = (
            f'=IFERROR(IF(ISBLANK(INDIRECT({_ref}&ADDRESS(ROW(),2))),"",INDIRECT({_ref}&ADDRESS(ROW(),2))),"")'
        )
        ws.cell(row=row, column=2).number_format = ACCOUNTING_FORMAT
        ws.cell(row=row, column=3).value = f'=IFERROR(B{row}/$B$4,"")'
        ws.cell(row=row, column=3).number_format = PERCENT_FORMAT
        ws.cell(row=row, column=4).value = (
            f'=IFERROR(IF(ISBLANK(INDIRECT({_ref}&ADDRESS(1,(ROW()-ROW($D$6))*2+1+2))),"",INDIRECT({_ref}&ADDRESS(1,(ROW()-ROW($D$6))*2+1+2))),"")'
        )
        ws.cell(row=row, column=5).value = (
            f'=IFERROR(IF(ISBLANK(INDIRECT({_ref}&ADDRESS(2,(ROW()-ROW($D$6))*2+1+1+2))),"",INDIRECT({_ref}&ADDRESS(2,(ROW()-ROW($D$6))*2+1+1+2))),"")'
        )
        ws.cell(row=row, column=5).number_format = ACCOUNTING_FORMAT
        ws.cell(row=row, column=6).value = f'=IFERROR(E{row}/$B$4,"")'
        ws.cell(row=row, column=6).number_format = PERCENT_FORMAT
        ws.cell(row=row, column=7).value = (
            f'=IFERROR(IF(ISBLANK(INDIRECT({_ref}&ADDRESS(3,(ROW()-ROW($D$6))*2+1+1+2))),"",INDIRECT({_ref}&ADDRESS(3,(ROW()-ROW($D$6))*2+1+1+2))),"")'
        )
        ws.cell(row=row, column=7).number_format = ACCOUNTING_FORMAT
        ws.cell(row=row, column=8).value = f'=IFERROR(G{row}/$B$4,"")'
        ws.cell(row=row, column=8).number_format = PERCENT_FORMAT

    # --- Column widths (prevents ##### for wide accounting-format numbers) ---
    for col_letter, width in SUMMARY_COLUMN_WIDTHS.items():
        ws.column_dimensions[col_letter].width = width

    # --- Borders ---
    # Vertical dividers: right of col C (Income | Expenses) and right of col H
    for row in range(1, 106):
        _update_cell_border(ws.cell(row=row, column=3), right=THIN)
        _update_cell_border(ws.cell(row=row, column=8), right=THIN)

    # Sub-dividers in header rows
    for row in [1, 2]:
        _update_cell_border(ws.cell(row=row, column=4), right=THIN)  # right of D1:D2
    _update_cell_border(ws.cell(row=2, column=6), right=THIN)       # right of F2
    _update_cell_border(ws.cell(row=4, column=6), right=THIN)       # right of F4

    # Top and bottom borders on section-label row, totals row, column-header row
    for row in [3, 4, 5]:
        for col in range(1, 9):
            _update_cell_border(ws.cell(row=row, column=col), top=THIN, bottom=THIN)


def _write_category_pair(ws, col_pair_idx, cat, cat_df, bold, center):
    """Write one category column pair (rows 1–5 structure + transactions) onto ws."""
    desc_col = col_pair_idx * 2 + 1
    amt_col = col_pair_idx * 2 + 2
    amt_col_letter = get_column_letter(amt_col)
    desc_col_letter = get_column_letter(desc_col)

    # Set column widths for this pair (prevents ##### for accounting numbers)
    ws.column_dimensions[desc_col_letter].width = RECEIPTS_DESC_COL_WIDTH
    ws.column_dimensions[amt_col_letter].width = RECEIPTS_AMT_COL_WIDTH

    # Row 1: category name merged across both columns (bold, centered)
    ws.cell(row=1, column=desc_col).value = cat
    ws.cell(row=1, column=desc_col).font = bold
    ws.cell(row=1, column=desc_col).alignment = center
    ws.merge_cells(
        start_row=1, start_column=desc_col,
        end_row=1, end_column=amt_col,
    )

    # Row 2: "Total" | SUM formula (accounting format)
    ws.cell(row=2, column=desc_col).value = "Total"
    ws.cell(row=2, column=desc_col).font = bold
    ws.cell(row=2, column=desc_col).alignment = center
    ws.cell(row=2, column=amt_col).value = (
        f"=SUM({amt_col_letter}6:{amt_col_letter}1001)"
    )
    ws.cell(row=2, column=amt_col).number_format = ACCOUNTING_FORMAT

    # Row 3: "Budget" | budgeted amount (accounting format)
    ws.cell(row=3, column=desc_col).value = "Budget"
    ws.cell(row=3, column=desc_col).font = bold
    ws.cell(row=3, column=desc_col).alignment = center
    ws.cell(row=3, column=amt_col).value = 0
    ws.cell(row=3, column=amt_col).number_format = ACCOUNTING_FORMAT

    # Row 4: empty (spacer)

    # Row 5: "Note" | "Amount" column headers (both bold, centered)
    ws.cell(row=5, column=desc_col).value = "Note"
    ws.cell(row=5, column=desc_col).font = bold
    ws.cell(row=5, column=desc_col).alignment = center
    ws.cell(row=5, column=amt_col).value = "Amount"
    ws.cell(row=5, column=amt_col).font = bold
    ws.cell(row=5, column=amt_col).alignment = center

    # Rows 6+: transactions (accounting format on amount column)
    for tx_idx, tx_row in cat_df.iterrows():
        excel_row = 6 + tx_idx
        ws.cell(row=excel_row, column=desc_col).value = tx_row["description"]
        ws.cell(row=excel_row, column=amt_col).value = tx_row["amount"]
        ws.cell(row=excel_row, column=amt_col).number_format = ACCOUNTING_FORMAT


def _apply_receipts_borders(ws, num_pairs):
    """
    Apply borders to a receipts sheet after all category pairs have been written.

    - Bottom border on the full width of rows 1, 4, and 5.
    - Right border on each category's amount column (col 2, 4, 6, …) for rows 1–1001.
    """
    last_col = num_pairs * 2

    # Bottom borders: category header row, spacer row (above Note/Amount), headers row
    for row in [1, 4, 5]:
        for col in range(1, last_col + 1):
            _update_cell_border(ws.cell(row=row, column=col), bottom=THIN)

    # Right border on each amount column to visually separate categories
    for pair_idx in range(num_pairs):
        amt_col = pair_idx * 2 + 2
        for row in range(1, 1002):
            _update_cell_border(ws.cell(row=row, column=amt_col), right=THIN)


def export_to_excel(input_file_path, output_file_path="budget_export.xlsx"):
    """
    Convert a CSV or JSON file (as produced by parse_receipts.py) back into
    the legacy Excel spreadsheet format.

    Input columns expected: date, category, description, amount, isPositive.
    The isPositive column determines whether a transaction is income (True) or
    an expense (False).  All income transactions are grouped into a single
    "Income" category that always occupies the first column pair on each sheet.

    The date column may use any format pandas can parse, including
    "M/D/YYYY" (e.g. "3/1/2026"), "YYYY-MM-DD" (e.g. "2026-03-01"),
    or datetime/Timestamp objects that pandas produces when reading JSON.

    Output Excel structure:
    - Sheet 1: "Summary" with income/expense overview and formulas
    - One additional sheet per month, named "M YYYY Receipts" (e.g. "3 2026 Receipts"),
      ordered latest-month-first.  Each category occupies a pair of columns:
      - Column pair 1: "Income" (all isPositive=True transactions)
      - Remaining pairs: one per unique expense category (isPositive=False)
      Each pair follows the structure:
      - Row 1: category name, merged across both columns, bold, centered
      - Row 2: "Total" (bold) | =SUM formula for the amount column
      - Row 3: "Budget" (bold) | budgeted amount (0 when not supplied)
      - Row 4: empty
      - Row 5: "Note" (bold) | "Amount" (bold)  — column headers
      - Row 6+: description (left col) / amount (right col) per transaction

    Parameters
    ----------
    input_file_path : str
        Path to the input CSV (.csv) or JSON (.json) file.
    output_file_path : str
        Path for the generated Excel workbook. Defaults to "budget_export.xlsx".

    Returns
    -------
    str
        The resolved path to the written Excel file.
    """
    ext = os.path.splitext(input_file_path)[1].lower()
    if ext == ".csv":
        df = pd.read_csv(input_file_path)
    elif ext == ".json":
        df = pd.read_json(input_file_path)
    else:
        raise ValueError(f"Unsupported file type: '{ext}'. Use .csv or .json.")

    required_columns = {"date", "category", "description", "amount"}
    missing = required_columns - set(df.columns)
    if missing:
        raise ValueError(f"Input file is missing required columns: {missing}")

    # Parse (month, year) from dates.
    # Accepts datetime/Timestamp objects as well as any date string that
    # pandas can parse (e.g. "M/D/YYYY", "YYYY-MM-DD", "MM-DD-YYYY", …).
    def _extract_month_year(date_val):
        if hasattr(date_val, "month") and hasattr(date_val, "year"):
            return int(date_val.month), int(date_val.year)
        try:
            ts = pd.to_datetime(str(date_val))
            return ts.month, ts.year
        except Exception:
            raise ValueError(
                f"Unrecognized date format '{date_val}'. "
                "Supported formats include 'M/D/YYYY' and 'YYYY-MM-DD'."
            )

    df["_month"], df["_year"] = zip(*df["date"].apply(_extract_month_year))

    bold = Font(bold=True)
    center = Alignment(horizontal="center")

    # One sheet per month, ordered in descending order (latest month first)
    month_groups = df.groupby(["_year", "_month"], sort=True)
    sorted_months = sorted(
        month_groups, key=lambda x: (x[0][0], x[0][1]), reverse=True
    )

    # Determine most recent month for the Summary tab
    if sorted_months:
        (most_recent_year, most_recent_month) = sorted_months[0][0]
    else:
        today = datetime.date.today()
        most_recent_year, most_recent_month = today.year, today.month

    wb = Workbook()
    ws_summary = wb.active
    ws_summary.title = "Summary"
    _write_summary_sheet(ws_summary, most_recent_month, most_recent_year)

    for (year, month), group in sorted_months:
        sheet_name = f"{month} {year} Receipts"
        ws = wb.create_sheet(title=sheet_name)

        # Sort transactions by date ascending within the month group
        group = group.copy()
        group["_date_parsed"] = pd.to_datetime(group["date"])
        group = group.sort_values("_date_parsed").drop(columns=["_date_parsed"])
        group = group.reset_index(drop=True)

        # Separate income (isPositive=True) and expense transactions.
        # If no isPositive column is present, treat all as expenses.
        if "isPositive" in group.columns:
            income_mask = group["isPositive"].apply(_to_bool)
        else:
            income_mask = pd.Series([False] * len(group), index=group.index)

        income_df = group[income_mask].reset_index(drop=True)
        expense_df = group[~income_mask].reset_index(drop=True)

        # Expense categories: apply priority ordering (PRIORITY_CATS first,
        # then any remaining categories in first-seen insertion order).
        _priority_map = {cat: i for i, cat in enumerate(PRIORITY_CATS)}
        seen_order = list(dict.fromkeys(expense_df["category"].tolist()))
        expense_cats = sorted(
            seen_order,
            key=lambda c: (_priority_map.get(c.lower(), len(PRIORITY_CATS)), seen_order.index(c)),
        )

        # "Income" is always the first column pair; expense categories follow
        col_pairs = [("Income", income_df)] + [
            (cat, expense_df[expense_df["category"] == cat].reset_index(drop=True))
            for cat in expense_cats
        ]

        for col_pair_idx, (cat, cat_df) in enumerate(col_pairs):
            _write_category_pair(ws, col_pair_idx, cat, cat_df, bold, center)

        _apply_receipts_borders(ws, len(col_pairs))

    wb.save(output_file_path)
    return os.path.realpath(output_file_path)


def main():
    parser = argparse.ArgumentParser(
        description=(
            "Convert a CSV or JSON budget file into the legacy Excel spreadsheet format. "
            "This is the inverse of parse_receipts.py."
        )
    )
    parser.add_argument(
        "input",
        help="Path to the input CSV or JSON file.",
    )
    parser.add_argument(
        "-o",
        "--output",
        default="budget_export.xlsx",
        help="Path for the output Excel file (default: budget_export.xlsx).",
    )
    args = parser.parse_args()

    output_path = export_to_excel(args.input, args.output)
    print(f"Exported to: {output_path}")


if __name__ == "__main__":
    main()
