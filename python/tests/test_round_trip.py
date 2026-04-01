"""Round-trip tests for parse_receipts.py and export_to_excel.py.

Verifies that data written by export_to_excel can be read back by
parse_receipts and produces an identical dataset.
"""

import datetime
import os
import tempfile

import pytest

# Require pandas for these round-trip tests; skip the module if it's missing.
pytest.importorskip("pandas")
import pandas as pd
from openpyxl import load_workbook

from export_to_excel import export_to_excel
from parse_receipts import parse_receipts


SAMPLE_ROWS = [
    {
        "date": "3/1/2026",
        "category": "groceries",
        "description": "Whole Foods",
        "amount": 45.50,
        "isPositive": False,
    },
    {
        "date": "3/1/2026",
        "category": "groceries",
        "description": "Trader Joe's",
        "amount": 78.25,
        "isPositive": False,
    },
    {
        "date": "3/1/2026",
        "category": "utilities",
        "description": "Electric Bill",
        "amount": 120.00,
        "isPositive": False,
    },
    {
        "date": "3/1/2026",
        "category": "income",
        "description": "Salary",
        "amount": 3000.00,
        "isPositive": True,
    },
    {
        "date": "4/1/2026",
        "category": "groceries",
        "description": "Costco",
        "amount": 200.00,
        "isPositive": False,
    },
    {
        "date": "4/1/2026",
        "category": "utilities",
        "description": "Water Bill",
        "amount": 60.00,
        "isPositive": False,
    },
]


SAMPLE_ROWS_ISO_DATES = [
    {
        "date": "2026-03-16",
        "category": "car",
        "description": "brakes",
        "amount": 87.3,
        "isPositive": False,
    },
    {
        "date": "2026-03-01",
        "category": "car",
        "description": "gas",
        "amount": 57.76,
        "isPositive": False,
    },
    {
        "date": "2026-03-01",
        "category": "rent",
        "description": "Rent",
        "amount": 1900.0,
        "isPositive": False,
    },
]


def _make_input_df():
    return pd.DataFrame(SAMPLE_ROWS)


def _make_iso_date_df():
    return pd.DataFrame(SAMPLE_ROWS_ISO_DATES)


def _normalize_round_trip_df(df: pd.DataFrame) -> pd.DataFrame:
    """Sort by stable keys and reset index to make comparisons order-insensitive."""
    sort_cols = ["date", "category", "description", "amount", "isPositive"]
    existing_cols = [c for c in sort_cols if c in df.columns]
    return df.sort_values(by=existing_cols).reset_index(drop=True)


# ---------------------------------------------------------------------------
# Round-trip via CSV
# ---------------------------------------------------------------------------


def test_round_trip_csv():
    """CSV -> export_to_excel -> parse_receipts reproduces the original data."""
    original = _make_input_df()

    with tempfile.TemporaryDirectory() as tmp:
        csv_path = os.path.join(tmp, "input.csv")
        xlsx_path = os.path.join(tmp, "output.xlsx")

        original.to_csv(csv_path, index=False)
        export_to_excel(csv_path, xlsx_path)
        result = parse_receipts(xlsx_path)

    result = _normalize_round_trip_df(result)
    original = _normalize_round_trip_df(original)

    assert list(result["date"]) == list(original["date"])
    assert list(result["category"]) == list(original["category"])
    assert list(result["description"]) == list(original["description"])
    assert list(result["amount"]) == pytest.approx(list(original["amount"]))
    assert list(result["isPositive"]) == list(original["isPositive"])


# ---------------------------------------------------------------------------
# Round-trip via JSON
# ---------------------------------------------------------------------------


def test_round_trip_json():
    """JSON -> export_to_excel -> parse_receipts reproduces the original data."""
    original = _make_input_df()

    with tempfile.TemporaryDirectory() as tmp:
        json_path = os.path.join(tmp, "input.json")
        xlsx_path = os.path.join(tmp, "output.xlsx")

        original.to_json(json_path, orient="records")
        export_to_excel(json_path, xlsx_path)
        result = parse_receipts(xlsx_path)

    result = _normalize_round_trip_df(result)
    original = _normalize_round_trip_df(original)

    assert list(result["date"]) == list(original["date"])
    assert list(result["category"]) == list(original["category"])
    assert list(result["description"]) == list(original["description"])
    assert list(result["amount"]) == pytest.approx(list(original["amount"]))
    assert list(result["isPositive"]) == list(original["isPositive"])


# ---------------------------------------------------------------------------
# Round-trip with ISO-format dates (YYYY-MM-DD)
# ---------------------------------------------------------------------------


def test_round_trip_iso_dates():
    """CSV with YYYY-MM-DD dates -> export_to_excel -> parse_receipts works correctly."""
    original = _make_iso_date_df()
    # parse_receipts always outputs dates as "M/1/YYYY" from the sheet name
    expected_dates = ["3/1/2026", "3/1/2026", "3/1/2026"]

    with tempfile.TemporaryDirectory() as tmp:
        csv_path = os.path.join(tmp, "input.csv")
        xlsx_path = os.path.join(tmp, "output.xlsx")

        original.to_csv(csv_path, index=False)
        export_to_excel(csv_path, xlsx_path)
        result = parse_receipts(xlsx_path)

    assert sorted(result["date"].tolist()) == sorted(expected_dates)

    # Compare category/description/amount independent of row order (category
    # priority ordering may differ from original insertion order).
    sort_keys = ["category", "description", "amount"]
    result_sorted = result.sort_values(by=sort_keys).reset_index(drop=True)
    original_sorted = original.sort_values(by=sort_keys).reset_index(drop=True)

    assert list(result_sorted["category"]) == list(original_sorted["category"])
    assert list(result_sorted["description"]) == list(original_sorted["description"])
    assert list(result_sorted["amount"]) == pytest.approx(list(original_sorted["amount"]))


# ---------------------------------------------------------------------------
# Excel structure
# ---------------------------------------------------------------------------


def test_excel_sheet_names():
    """Exported workbook has a Summary sheet plus one sheet per month in descending order."""
    original = _make_input_df()

    with tempfile.TemporaryDirectory() as tmp:
        csv_path = os.path.join(tmp, "input.csv")
        xlsx_path = os.path.join(tmp, "output.xlsx")

        original.to_csv(csv_path, index=False)
        export_to_excel(csv_path, xlsx_path)

        xl = pd.ExcelFile(xlsx_path)
        assert xl.sheet_names[0] == "Summary"
        assert "3 2026 Receipts" in xl.sheet_names
        assert "4 2026 Receipts" in xl.sheet_names
        # Latest month should appear first after Summary (descending order)
        month_sheets = [s for s in xl.sheet_names if s != "Summary"]
        assert month_sheets[0] == "4 2026 Receipts"
        assert month_sheets[1] == "3 2026 Receipts"


def test_summary_sheet_layout():
    """Summary tab has the full layout: Date/Cash Flow headers, Income/Expenses sections."""
    original = _make_input_df()

    with tempfile.TemporaryDirectory() as tmp:
        csv_path = os.path.join(tmp, "input.csv")
        xlsx_path = os.path.join(tmp, "output.xlsx")

        original.to_csv(csv_path, index=False)
        export_to_excel(csv_path, xlsx_path)

        wb = load_workbook(xlsx_path, data_only=False)

    ws = wb["Summary"]

    # A1 is merged and contains "Date"
    assert ws.cell(row=1, column=1).value == "Date"
    assert ws.cell(row=1, column=1).font.bold

    # B1 is merged and contains a date (for the most-recent month)
    assert isinstance(ws.cell(row=1, column=2).value, datetime.date)

    # D1 is merged and contains "Cash Flow"
    assert ws.cell(row=1, column=4).value == "Cash Flow"
    assert ws.cell(row=1, column=4).font.bold

    # Row 1 column headers
    assert ws.cell(row=1, column=5).value == "Amount"
    assert ws.cell(row=1, column=6).value == "Fraction"
    assert ws.cell(row=1, column=7).value == "Budget"
    assert ws.cell(row=1, column=8).value == "Fraction"

    # Row 2: cash-flow formulas in E–H
    assert ws.cell(row=2, column=5).value == "=$B$4-E4"
    assert ws.cell(row=2, column=7).value == "=$B$4-G4"

    # Row 3: section headers
    assert ws.cell(row=3, column=1).value == "Income"
    assert ws.cell(row=3, column=4).value == "Expenses"

    # Row 4: totals row
    assert ws.cell(row=4, column=1).value == "Total"
    assert ws.cell(row=4, column=2).value == "=SUM(B6:B1048576)"
    assert ws.cell(row=4, column=4).value == "Total"
    assert ws.cell(row=4, column=5).value == "=SUM(E6:E102)"

    # Row 5: column headers
    assert ws.cell(row=5, column=1).value == "Source"
    assert ws.cell(row=5, column=2).value == "Amount"
    assert ws.cell(row=5, column=4).value == "Item"

    # Row 6+: formulas are present
    assert ws.cell(row=6, column=1).value is not None
    assert ws.cell(row=6, column=4).value is not None

    # Verify merged cells: A1:A2, B1:C2, D1:D2, A3:C3, D3:H3
    merged_strs = {str(r) for r in ws.merged_cells.ranges}
    assert "A1:A2" in merged_strs
    assert "B1:C2" in merged_strs
    assert "D1:D2" in merged_strs
    assert "A3:C3" in merged_strs
    assert "D3:H3" in merged_strs


def test_excel_category_row():
    """Row 1 of each data sheet contains category names at every other column."""
    original = _make_input_df()

    with tempfile.TemporaryDirectory() as tmp:
        csv_path = os.path.join(tmp, "input.csv")
        xlsx_path = os.path.join(tmp, "output.xlsx")

        original.to_csv(csv_path, index=False)
        export_to_excel(csv_path, xlsx_path)

        xl = pd.ExcelFile(xlsx_path)
        sheet = xl.parse("3 2026 Receipts", header=None)

    category_row = sheet.iloc[0]
    # Even columns hold category names; odd columns are empty (merged cell spans both)
    even_values = [v for i, v in enumerate(category_row) if i % 2 == 0]
    odd_values = [v for i, v in enumerate(category_row) if i % 2 == 1]

    assert all(pd.notna(v) for v in even_values)
    assert all(pd.isna(v) for v in odd_values)

    # "Income" is always the first category
    assert even_values[0] == "Income"


def test_excel_structure_rows():
    """Verifies Total, Budget, and Note/Amount header rows in each data sheet."""
    original = _make_input_df()

    with tempfile.TemporaryDirectory() as tmp:
        csv_path = os.path.join(tmp, "input.csv")
        xlsx_path = os.path.join(tmp, "output.xlsx")

        original.to_csv(csv_path, index=False)
        export_to_excel(csv_path, xlsx_path)

        xl = pd.ExcelFile(xlsx_path)
        sheet = xl.parse("3 2026 Receipts", header=None)

    # Row 2 (index 1): description columns hold "Total"
    total_row = sheet.iloc[1]
    desc_vals = [total_row.iloc[i] for i in range(0, len(total_row), 2)]
    assert all(v == "Total" for v in desc_vals)

    # Row 3 (index 2): description columns hold "Budget"
    budget_row = sheet.iloc[2]
    budget_desc_vals = [budget_row.iloc[i] for i in range(0, len(budget_row), 2)]
    assert all(v == "Budget" for v in budget_desc_vals)

    # Row 5 (index 4): "Note" in description columns, "Amount" in amount columns
    note_row = sheet.iloc[4]
    note_vals = [note_row.iloc[i] for i in range(0, len(note_row), 2)]
    amount_header_vals = [note_row.iloc[i] for i in range(1, len(note_row), 2)]
    assert all(v == "Note" for v in note_vals)
    assert all(v == "Amount" for v in amount_header_vals)

    # Transactions start at row 6 (index 5)
    first_tx_desc = sheet.iloc[5, 0]
    assert pd.notna(first_tx_desc)


def test_excel_bold_formatting():
    """Category names, Total, Budget, and Amount cells are bold."""
    original = _make_input_df()

    with tempfile.TemporaryDirectory() as tmp:
        csv_path = os.path.join(tmp, "input.csv")
        xlsx_path = os.path.join(tmp, "output.xlsx")

        original.to_csv(csv_path, index=False)
        export_to_excel(csv_path, xlsx_path)

        wb = load_workbook(xlsx_path)

    ws = wb["3 2026 Receipts"]

    # Row 1, col 1 (A1): category name — bold and centered
    assert ws.cell(row=1, column=1).font.bold
    assert ws.cell(row=1, column=1).alignment.horizontal == "center"

    # Row 2, col 1 (A2): "Total" — bold
    assert ws.cell(row=2, column=1).font.bold

    # Row 3, col 1 (A3): "Budget" — bold
    assert ws.cell(row=3, column=1).font.bold

    # Row 5, col 1 (A5): "Note" — bold
    assert ws.cell(row=5, column=1).font.bold

    # Row 5, col 2 (B5): "Amount" — bold
    assert ws.cell(row=5, column=2).font.bold


def test_excel_total_formula():
    """Row 2 amount cells contain a SUM formula referencing row 6 onward."""
    original = _make_input_df()

    with tempfile.TemporaryDirectory() as tmp:
        csv_path = os.path.join(tmp, "input.csv")
        xlsx_path = os.path.join(tmp, "output.xlsx")

        original.to_csv(csv_path, index=False)
        export_to_excel(csv_path, xlsx_path)

        # data_only=False preserves formula strings
        wb = load_workbook(xlsx_path, data_only=False)

    ws = wb["3 2026 Receipts"]
    # B2: SUM formula for first category
    formula = ws.cell(row=2, column=2).value
    assert formula is not None
    assert formula.startswith("=SUM(B6:")


def test_excel_merged_category_header():
    """Category header cells are merged across both description and amount columns."""
    original = _make_input_df()

    with tempfile.TemporaryDirectory() as tmp:
        csv_path = os.path.join(tmp, "input.csv")
        xlsx_path = os.path.join(tmp, "output.xlsx")

        original.to_csv(csv_path, index=False)
        export_to_excel(csv_path, xlsx_path)

        wb = load_workbook(xlsx_path)

    ws = wb["3 2026 Receipts"]
    # Collect all merged ranges in row 1
    row1_merges = [
        r for r in ws.merged_cells.ranges if r.min_row == 1 and r.max_row == 1
    ]
    # Expected: 1 "Income" pair + one pair per unique expense category in March
    input_df = _make_input_df()
    march_rows = input_df[input_df["date"].str.startswith("3/")]
    expense_cats = set(
        march_rows[~march_rows["isPositive"]]["category"].tolist()
    )
    expected_pairs = 1 + len(expense_cats)  # Income + expense categories
    assert len(row1_merges) == expected_pairs


def test_invalid_file_type_raises():
    """export_to_excel raises ValueError for unsupported file extensions."""
    with tempfile.TemporaryDirectory() as tmp:
        with pytest.raises(ValueError, match="Unsupported file type"):
            export_to_excel("data.txt", os.path.join(tmp, "out.xlsx"))


def test_missing_columns_raises():
    """export_to_excel raises ValueError when required columns are absent."""
    with tempfile.TemporaryDirectory() as tmp:
        csv_path = os.path.join(tmp, "bad.csv")
        pd.DataFrame({"date": ["3/1/2026"], "amount": [10.0]}).to_csv(
            csv_path, index=False
        )
        with pytest.raises(ValueError, match="missing required columns"):
            export_to_excel(csv_path, os.path.join(tmp, "out.xlsx"))
